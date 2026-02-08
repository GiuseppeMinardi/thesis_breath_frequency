"""
Calculate work hours and associated costs from a CSV log file.

This module reads a CSV file containing work session start and end times,
calculates the duration of each session, computes the associated cost based
on an hourly rate, and exports the results to a formatted Excel file.
"""

import argparse
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

import pandas as pd


def load_csv(file_path: Path) -> pd.DataFrame:
    """Load CSV data into a DataFrame."""
    return pd.read_csv(file_path)


def process_data(
    df_: pd.DataFrame,
    date_format: str = "%d/%m/%Y_%H:%M",
    hourly_fee: float = 25.0,
) -> pd.DataFrame:
    """Process the DataFrame to calculate durations and costs."""
    df_ = df_.assign(
        start=lambda df_: pd.to_datetime(df_.start, format=date_format),
        end=lambda df_: pd.to_datetime(df_.end, format=date_format),
    )
    df_["duration_hrs"] = df_["end"].sub(df_["start"]).dt.total_seconds().div(3600)
    df_["cost_eur"] = df_["duration_hrs"] * hourly_fee
    return df_


def modify_excel(file_path: Path) -> None:
    """Apply styling and formatting to the generated Excel file."""
    wb = load_workbook(file_path)
    ws = wb.active
    last_row = ws.max_row

    # Styles
    header_fill = PatternFill(
        start_color="1F4E78", end_color="1F4E78", fill_type="solid"
    )
    summary_fill = PatternFill(
        start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"
    )
    white_font = Font(color="FFFFFF", bold=True)
    bold_font = Font(bold=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # 1. Format Main Table Headers
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # 2. Format Data Body
    for row in ws.iter_rows(min_row=2, max_row=last_row):
        row[2].number_format = '0.00 "hrs"'  # Duration (Col C)
        row[3].number_format = "€ #,##0.00"  # Cost (Col D)

        for i in range(0, 5):  # Center align Duration and Cost
            row[i].alignment = Alignment(vertical="center")
        row[4].alignment = Alignment(wrap_text=True, vertical="center")  # Note (Col E)
        for cell in row:
            cell.border = thin_border

    # 3. Dedicated Summary Section (Top Right)
    total_hours_col = 8  # H
    total_cost_col = 9  # I

    ws.cell(row=1, column=total_hours_col, value="TOTAL HOURS").font = bold_font
    ws.cell(row=1, column=total_cost_col, value="TOTAL COST").font = bold_font
    ws.cell(row=2, column=7, value="TOTAL DUE:").font = bold_font
    ws.cell(row=2, column=7).alignment = Alignment(
        horizontal="right", vertical="center"
    )

    # Corrected Formulas: Summing C (Duration) and D (Cost)
    total_hrs_cell = ws.cell(
        row=2, column=total_hours_col, value=f"=SUM(C2:C{last_row})"
    )
    total_cost_cell = ws.cell(
        row=2, column=total_cost_col, value=f"=SUM(D2:D{last_row})"
    )

    # Styling Summary Area (G1:I2)
    for col in range(7, 10):
        for r in range(1, 3):
            cell = ws.cell(row=r, column=col)
            cell.border = thin_border
            if r == 2 and col >= 8:
                cell.fill = summary_fill
                cell.font = bold_font

    total_hrs_cell.number_format = '0.00 "hrs"'
    total_hrs_cell.alignment = Alignment(horizontal="center", vertical="center")
    total_cost_cell.number_format = "€ #,##0.00"
    total_cost_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Adjust Column Widths
    widths = {"A": 20, "B": 20, "C": 10, "D": 10, "E": 30, "G": 15, "H": 15, "I": 15}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    wb.save(file_path)


def main(
    input_csv_path: Path, output_excel_path: Path, date_format: str, hourly_fee: float
) -> None:
    """Execute the data processing and Excel generation."""
    dataset = load_csv(input_csv_path)
    processed_data = process_data(
        dataset, date_format=date_format, hourly_fee=hourly_fee
    )
    processed_data = processed_data[
        ["start", "end", "duration_hrs", "cost_eur", "note"]
    ].rename(
        columns={
            "start": "Start Time",
            "end": "End Time",
            "duration_hrs": "Duration (hrs)",
            "cost_eur": "Cost (€)",
            "note": "Note",
        }
    )
    processed_data.to_excel(output_excel_path, index=False)
    modify_excel(output_excel_path)
    print(f"Data processed and saved to {output_excel_path}")


if __name__ == "__main__":
    path_file: Path = Path(__file__).parent.joinpath("ore_lavorate.csv")
    output_file: Path = Path(__file__).parent.joinpath("ore_lavorate_calcolate.xlsx")
    hourly_fee: float = 25.0

    parser = argparse.ArgumentParser(
        description="Calculate work hours and costs from a CSV log."
    )
    parser.add_argument(
        "--input_csv",
        type=Path,
        default=path_file,
        help="Path to the input CSV file containing work session logs.",
    )
    parser.add_argument(
        "-o",
        "--output_excel",
        type=Path,
        default=output_file,
        help="Path to the output Excel file where results will be saved.",
    )

    parser.add_argument(
        "-f",
        "--hourly_fee",
        type=float,
        default=25.0,
        help="Hourly fee in euros to calculate costs (default: 25.0).",
    )

    parser.add_argument(
        "-df",
        "--date_format",
        type=str,
        default="%d/%m/%Y_%H:%M",
        help="Date format to parse input CSV dates (default: %Y-%m-%d).",
    )

    parser_args = parser.parse_args()
    path_file = parser_args.input_csv
    output_file = parser_args.output_excel
    hourly_fee = parser_args.hourly_fee
    date_format = parser_args.date_format

    if not path_file.exists():
        path_file.touch()  # Create an empty file if it doesn't exist
        with path_file.open("w") as f:
            f.write("start,end,note\n")  # Write header for the CSV
        sys.exit(
            f"Created empty CSV at {path_file}. Please populate it with data and rerun the script."
        )
    main(
        input_csv_path=path_file,
        output_excel_path=output_file,
        date_format=date_format,
        hourly_fee=hourly_fee,
    )
