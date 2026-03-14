import openpyxl

import pandas as pd


def normalize_columns(df_: pd.DataFrame) -> pd.DataFrame:
    df_.columns = [col.strip().lower().replace(" ", "_") for col in df_.columns]
    return df_


def convert_time_to_seconds(time_str: str) -> int:
    if isinstance(time_str, str):
        hours, minutes, seconds = time_str.split(":")
        minutes = int(minutes) + int(hours) * 60
        seconds = int(seconds) + minutes * 60
        return seconds
    else:
        raise ValueError(f"Expected a string in the format 'HH:MM:SS', got {time_str}")


def correct_df_dtypes(df: pd.DataFrame) -> pd.DataFrame:
    # 1. Strip whitespace from all string columns
    str_cols = df.select_dtypes(include=["object", "string"]).columns
    for col in str_cols:
        df[col] = df[col].astype(str).str.strip()

    # 2. Attempt to convert string columns to numeric where possible
    for col in str_cols:
        # check if the column is primarily digits/decimals
        # errors='ignore' keeps it as string if it's not a number
        converted = pd.to_numeric(df[col], errors="coerce")

        # If the conversion didn't result in all NaNs (meaning it was actually numbers)
        if not converted.isna().all():
            df[col] = converted

    return df


def rename_columns(df: pd.DataFrame) -> pd.DataFrame:
    rename_mapping = {
        "work_(watts)": "work_watts",
        "vo2_(ml/kg/min)": "vo2_ml_per_kg_min",
        "vo2_(ml/min)": "vo2_ml_per_min",
        "vco2_(ml/min)": "vco2_ml_per_min",
        "rer": "rer",
        "rr_(br/min)": "rr_br_per_min",
        "vt_btps_(l)": "vt_btps_l",
        "ve_btps_(l/min)": "ve_btps_l_per_min",
        "br_(%)": "breathing_reserve_pct",
        "hr_(bpm)": "hr_bpm",
        "hrr_(%)": "hrr_pct",
        "peto2_(mmhg)": "peto2_mmhg",
        "petco2_(mmhg)": "petco2_mmhg",
        "rr_(br/min)copy": "rr_br_per_min_copy",
        "vo2/pred(%)": "vo2_pred_pct",
        "ti/ttot": "ti_ttot_ratio",
        "ti_(sec)": "ti_sec",
        "te_(sec)": "te_sec",
        "ttot_(sec)": "ttot_sec",
        "msec": "rr_interval_msec",
        "msec_diff_quad": "rr_interval_diff_squared_msec",
        "rmssq": "rmssd_ms",
        "patient": "patient_id",
        "time_seconds": "time_seconds",
    }
    return df.rename(columns=rename_mapping)

def find_breakpoints(
    workbook: openpyxl.Workbook,
    sheet_name: str,
    skiprows: int,
    targets: list[str] | None = None,
    stop_when_all_found: bool = True,
    max_rows: int | None = None,
) -> dict[str, list[int]]:
    if targets is None:
        # Aggiunto "V02 MAX" con lo zero per matchare i refusi nei file grezzi
        targets = ["AT", "RC", "VO2 MAX", "V02 MAX"]

    # Normalizza i target (maiuscolo e senza spazi)
    normalized_to_label: dict[str, str] = {
        str(t).strip().upper(): str(t) for t in targets
    }

    sheet = workbook[sheet_name]
    # Each entry will be a list of dicts: {"index": int, "time_seconds": int|None, "work": float|None}
    found: dict[str, list[dict]] = {label: [] for label in normalized_to_label.values()}

    # CONTATORE CRITICO: tiene traccia di quante righe di testo abbiamo
    # già incontrato, per sottrarle dall'indice Pandas
    markers_found_count = 0

    for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if max_rows is not None and row_idx > max_rows:
            break

        if not row or row[0] is None:
            continue

        cell_value = str(row[0]).strip().upper()

        if cell_value in normalized_to_label:
            label = normalized_to_label[cell_value]

            # Calcolo dell'indice Pandas corretto:
            # -1 per passare a 0-based
            # -markers_found_count per compensare le righe target precedenti che verranno droppate
            pandas_idx = (row_idx - skiprows) - 1 - markers_found_count

            # Se hai "V02 MAX" come refuso, puoi forzarlo a raggrupparsi sotto "VO2 MAX"
            if label == "V02 MAX":
                label = "VO2 MAX"
                if label not in found:
                    found[label] = []

            # Try to read the data row following the marker to capture time and work.
            time_seconds = None
            work_val = None
            try:
                # The data row that corresponds to the pandas index is typically the
                # row immediately after the marker row in the raw sheet.
                data_row_number = row_idx + 1
                time_cell = sheet.cell(row=data_row_number, column=1).value
                work_cell = sheet.cell(row=data_row_number, column=2).value

                if time_cell is not None:
                    try:
                        time_seconds = convert_time_to_seconds(str(time_cell))
                    except Exception:
                        time_seconds = None

                if work_cell is not None:
                    try:
                        work_val = float(str(work_cell).strip())
                    except Exception:
                        work_val = None
            except Exception:
                time_seconds = None
                work_val = None

            found[label].append(
                {
                    "index": pandas_idx,
                    "time_seconds": time_seconds,
                    "work": work_val,
                }
            )
            markers_found_count += 1

        # Stop anticipato se tutti i target hanno almeno un valore
        if stop_when_all_found and all(found[lbl] for lbl in ["AT", "RC", "VO2 MAX"]):
            break

    # Rimuove le entry vuote
    return {k: v for k, v in found.items() if v}