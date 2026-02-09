"""Load and process Excel respiratory data for the thesis project.

This module provides helpers to read Excel files containing multiple
patient sheets, normalize and convert columns, and return a single
concatenated :class:`pandas.DataFrame` ready for downstream analysis.

Functions
---------
_get_skip_rows
    Return the number of rows to skip for known sheet names.
load_and_process_excel
    Load an Excel file, process each valid sheet and concatenate results.
"""

from pathlib import Path

from openpyxl import load_workbook

import pandas as pd

from ..__init__ import logger
from ._cleaning_utils import (
    convert_time_to_seconds,
    correct_df_dtypes,
    normalize_columns,
    rename_columns,
)

header = [
    "Time (min)",
    "Work (Watts)",
    "VO2 (mL/kg/min)",
    "VO2 (mL/min)",
    "VCO2 (mL/min)",
    "RER",
    "RR (br/min)",
    "Vt BTPS (L)",
    "VE BTPS (L/min)",
    "BR (%)",
    "HR (BPM)",
    "HRR (%)",
    "PETO2 (mmHg)",
    "PETCO2 (mmHg)",
    "RR (br/min) copy",
    "VO2/Pred (%)",
    "Ti/Ttot",
    "Ti (sec)",
    "Te (sec)",
    "Ttot (sec)",
    "msec",
    "msec_diff_quad",
    "RMSSQ",
]


def _get_skip_rows(sheet_name: str) -> int:
    """Get the number of rows to skip for a given sheet name."""
    skip_rows_map = {
        "FG": 3,
        "VR": 4,
        "CN": 4,
        "SG": 3,
        "AA": 4,
        "MS": 4,
        "GE": 4,
        "PT": 4,
        "CM": 3,
        "CN2": 4,
        # "SM": 2,  # TODO: controlla SM
        "CG": 3,
        "MP": 3,
        "IL": 4,
        "GR": 4,
        "VG": 4,
        "GP": 4,
        "CL": 4,
        "GM": 4,
        "SL": 4,
    }
    return skip_rows_map.get(sheet_name)


def load_and_process_excel(source_excel_path: Path) -> pd.DataFrame:
    """
    Load and process Excel file with multiple sheet tabs.

    Args:
        source_excel_path: Path to the Excel file to process.

    Returns
    -------
        A concatenated DataFrame with processed data from all valid sheets.
    """
    dataframes = []
    for sheet_name in load_workbook(source_excel_path, read_only=True).sheetnames:
        logger.info(f"Processing sheet: {sheet_name}")
        skip_to_row = _get_skip_rows(sheet_name)
        if skip_to_row is None:
            logger.warning(f"No skip row defined for sheet '{sheet_name}'. Skipping this sheet.")
            continue
        logger.debug(f"Processing sheet '{sheet_name}' with skip_to_row={skip_to_row}")

        dataframes.append(
            pd.read_excel(
                source_excel_path,
                sheet_name=sheet_name,
                skiprows=skip_to_row,
                usecols="A:W",
                names=header,
                dtype="string",
            )
            .pipe(normalize_columns)
            .dropna(thresh=20)
            .assign(
                patient=sheet_name,
                time_seconds=lambda df_: df_["time_(min)"].apply(convert_time_to_seconds),
            )
            .drop(columns=["time_(min)"])
            .pipe(correct_df_dtypes)
            .pipe(rename_columns)
        )
    logger.info(f"Processed {len(dataframes)} sheets from the Excel file.")
    return pd.concat(dataframes)